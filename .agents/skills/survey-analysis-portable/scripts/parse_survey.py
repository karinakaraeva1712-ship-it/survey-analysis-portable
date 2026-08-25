#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Parse a survey export (xlsx) into survey.json.

Usage:
  uv run --with openpyxl python parse_survey.py <input.xlsx> [output.json]

Output contract:
  {n, meta_columns, questions, options, scale_map, header, data}

This export format (report sheet of the survey platform):
  - Header row is the FIRST row; first cell == 'Answer ID'.
  - Column headers embed the question structure:
      '2. Choice – <question text>'            single-select (value = option text)
      '2. Choice, other answers – <text>'      free-text 'other' of single-select
      '5. Choice – <question>: <option text>'  multiple-select flag column
      '17. Question – <question text>'         open-ended text question
  - Multi-select cells hold 'TRUE'/'FALSE' strings.
  - There are no dictionary rows and no grid/NPS questions in this export.
"""
import json
import re
import sys
from collections import Counter

from openpyxl import load_workbook

Q_RE = re.compile(r"^(\d+)\. (Choice|Question)(?:, other answers)? – (.*)$")


def s(v):
    if v is None:
        return None
    return str(v).strip()


def parse_question_code(header_text):
    """Classify one question column header."""
    m = Q_RE.match(header_text)
    if not m:
        return None
    num, kind, text = m.group(1), m.group(2), m.group(3)
    q = f"Q_{num}"
    if kind == "Question":
        return {"q": q, "type": "open", "text": text}
    if ", other answers" in header_text:
        return {"q": q, "type": "other", "text": text}
    if ": " in text:
        qtext, opt = text.split(": ", 1)
        return {"q": q, "type": "multi", "text": qtext, "opt": opt}
    return {"q": q, "type": "single", "text": text}


def main():
    if len(sys.argv) < 2:
        raise SystemExit("Usage: parse_survey.py <input.xlsx> [output.json]")
    inp = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else "survey.json"

    wb = load_workbook(inp, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    hdr = rows[0]

    meta_columns = []
    header = {}
    questions = {}
    options = {}
    multi_counter = Counter()
    single_raw = {}
    other_cols = {}

    for j, cell in enumerate(hdr):
        if cell is None or s(cell) == "":
            continue
        t = s(cell)
        if j < 13 or t == "Answer ID":
            meta_columns.append(t)
            header[j] = f"meta_{j}"
            continue
        info = parse_question_code(t)
        if info is None:
            meta_columns.append(t)
            header[j] = f"meta_{j}"
            continue
        q = info["q"]
        typ = info["type"]
        if typ == "open":
            header[j] = q
            questions[q] = {"type": "Open text", "text": info["text"]}
        elif typ == "multi":
            idx = multi_counter[q]
            code = f"{q}_{chr(ord('a') + idx)}"
            multi_counter[q] += 1
            header[j] = code
            options[code] = info["opt"]
            if q not in questions:
                questions[q] = {"type": "Multiple select", "text": info["text"]}
        elif typ == "single":
            header[j] = q
            questions[q] = {"type": "Single select", "text": info["text"]}
            single_raw[q] = code if False else q
        elif typ == "other":
            code = f"{q}_other"
            header[j] = code
            questions[q] = {"type": questions.get(q, {}).get("type", "Single select"),
                            "text": questions.get(q, {}).get("text", info["text"])}
            other_cols[q] = code

    data = []
    for r in rows[1:]:
        rec = {}
        for j, code in header.items():
            if j < len(r):
                v = r[j]
                if v is None:
                    continue
                if isinstance(v, str):
                    sv = v.strip()
                    if sv == "":
                        continue
                    if sv in ("TRUE", "FALSE"):
                        v = 1 if sv == "TRUE" else 0
                    else:
                        v = sv
                rec[code] = v
        if rec:
            data.append(rec)

    # synthetic one-hot flags for single-select questions (for correlation analysis);
    # only answered respondents get a flag (absent = question not routed to them)
    for q, code in single_raw.items():
        vals = Counter(r.get(code) for r in data if r.get(code) is not None)
        order = sorted(vals, key=lambda x: (-vals[x], x))
        for i, val in enumerate(order):
            flag_code = f"{q}_{chr(ord('a') + i)}"
            options[flag_code] = val
            for r in data:
                if r.get(code) is not None:
                    r[flag_code] = 1 if r.get(code) == val else 0

    result = {
        "n": len(data),
        "meta_columns": meta_columns,
        "questions": questions,
        "options": options,
        "scale_map": {},
        "header": {str(j): c for j, c in sorted(header.items())},
        "data": data,
    }
    with open(out, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False)
    print(f"parsed: n={len(data)}, questions={len(questions)}, options={len(options)}, "
          f"multi_flag_cols={sum(multi_counter.values())}, "
          f"single_questions={len(single_raw)}, open={len([q for q,i in questions.items() if i['type']=='Open text'])}, "
          f"other_cols={len(other_cols)}, meta={len(meta_columns)}")


if __name__ == "__main__":
    main()
